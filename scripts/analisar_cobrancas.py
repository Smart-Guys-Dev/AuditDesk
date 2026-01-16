"""
Script para identificar procedimentos por período
Aba 1: Após 20/11/2025
Aba 2: Até 19/11/2025
"""

import os
import sys
from datetime import datetime
import lxml.etree as etree

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

NAMESPACES = {'ptu': 'http://ptu.unimed.coop.br/schemas/V3_0'}
DATA_CORTE = datetime(2025, 11, 20)


def parse_data(texto):
    if not texto:
        return None
    try:
        texto = texto.strip()
        if len(texto) >= 8 and texto[:8].isdigit():
            return datetime.strptime(texto[:8], "%Y%m%d")
        texto = texto[:10].replace("/", "-")
        return datetime.strptime(texto, "%Y-%m-%d")
    except:
        return None


def format_data(texto):
    dt = parse_data(texto)
    if dt:
        return dt.strftime("%d/%m/%Y")
    return texto or ""


def get_text(element, xpath, ns):
    vals = element.xpath(xpath, namespaces=ns)
    if vals:
        return vals[0].text if hasattr(vals[0], 'text') else str(vals[0])
    return ""


def analisar_xml(caminho_arquivo):
    try:
        tree = etree.parse(caminho_arquivo)
        root = tree.getroot()
    except:
        return [], []

    apos = []  # Após 20/11/2025
    ate = []   # Até 19/11/2025
    ns = NAMESPACES
    arquivo = os.path.basename(caminho_arquivo)
    
    def processar_guia(guia, tipo):
        nr_guia = get_text(guia, ".//ptu:nr_GuiaTissPrestador", ns) or get_text(guia, ".//ptu:nr_LotePrestador", ns)
        id_guia = get_text(guia, ".//ptu:nr_GuiaTissOperadora", ns)
        
        if tipo == "Consulta":
            cd_servico = get_text(guia, ".//ptu:cd_Servico", ns)
            ds_servico = get_text(guia, ".//ptu:ds_Servico", ns)
            dt_exec = get_text(guia, ".//ptu:dt_Atendimento", ns)
            
            dt = parse_data(dt_exec)
            if dt:
                registro = {
                    "arquivo": arquivo, "tipo": tipo, "nr_guia": nr_guia,
                    "id": id_guia, "cd_servico": cd_servico,
                    "ds_servico": ds_servico, "dt_execucao": format_data(dt_exec)
                }
                if dt > DATA_CORTE:
                    apos.append(registro)
                else:
                    ate.append(registro)
        else:
            for proc in guia.xpath(".//ptu:procedimentosExecutados", namespaces=ns):
                cd_servico = get_text(proc, ".//ptu:cd_Servico", ns)
                ds_servico = get_text(proc, ".//ptu:ds_Servico", ns)
                dt_exec = get_text(proc, ".//ptu:dt_Execucao", ns) or get_text(proc, ".//ptu:dt_Atendimento", ns)
                
                dt = parse_data(dt_exec)
                if dt:
                    registro = {
                        "arquivo": arquivo, "tipo": tipo, "nr_guia": nr_guia,
                        "id": id_guia, "cd_servico": cd_servico,
                        "ds_servico": ds_servico, "dt_execucao": format_data(dt_exec)
                    }
                    if dt > DATA_CORTE:
                        apos.append(registro)
                    else:
                        ate.append(registro)
    
    for guia in root.xpath(".//ptu:guiaConsulta", namespaces=ns):
        processar_guia(guia, "Consulta")
    for guia in root.xpath(".//ptu:guiaSADT", namespaces=ns):
        processar_guia(guia, "SADT")
    for guia in root.xpath(".//ptu:guiaInternacao", namespaces=ns):
        processar_guia(guia, "Internação")
    for guia in root.xpath(".//ptu:guiaHonorarios", namespaces=ns):
        processar_guia(guia, "Honorários")
    
    return apos, ate


def criar_planilha(ws, dados, titulo):
    ws.title = titulo
    headers = ["Arquivo", "Tipo Guia", "Nr. Guia", "ID", "Cód. Serviço", "Descrição Serviço", "Dt. Execução"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row, r in enumerate(dados, 2):
        ws.cell(row=row, column=1, value=r["arquivo"])
        ws.cell(row=row, column=2, value=r["tipo"])
        ws.cell(row=row, column=3, value=r["nr_guia"])
        ws.cell(row=row, column=4, value=r["id"])
        ws.cell(row=row, column=5, value=r["cd_servico"])
        ws.cell(row=row, column=6, value=r["ds_servico"])
        ws.cell(row=row, column=7, value=r["dt_execucao"])
    
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 14


def processar_e_exportar(caminho, arquivo_saida):
    arquivos = []
    if os.path.isfile(caminho):
        arquivos = [caminho]
    elif os.path.isdir(caminho):
        for raiz, _, files in os.walk(caminho):
            arquivos.extend(os.path.join(raiz, f) for f in files if f.lower().endswith(('.xml', '.051')))
    
    todos_apos = []
    todos_ate = []
    
    for arq in arquivos:
        apos, ate = analisar_xml(arq)
        todos_apos.extend(apos)
        todos_ate.extend(ate)
    
    # Criar Excel com 2 abas
    wb = Workbook()
    
    # Aba 1: Após 20/11/2025
    ws1 = wb.active
    criar_planilha(ws1, todos_apos, "Após 20-11-2025")
    
    # Aba 2: Até 19/11/2025
    ws2 = wb.create_sheet("Até 19-11-2025")
    criar_planilha(ws2, todos_ate, "Até 19-11-2025")
    
    wb.save(arquivo_saida)
    print(f"\n✅ Excel gerado: {arquivo_saida}")
    print(f"\n📋 RESUMO:")
    print(f"   - Aba 'Após 20-11-2025': {len(todos_apos)} procedimentos")
    print(f"   - Aba 'Até 19-11-2025': {len(todos_ate)} procedimentos")
    print(f"   - TOTAL: {len(todos_apos) + len(todos_ate)} procedimentos")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python analisar_cobrancas.py <caminho>")
        sys.exit(1)
    
    # Nome de saída baseado no arquivo de entrada
    base_name = os.path.basename(sys.argv[1]).replace('.xml', '').replace('.051', '')
    saida = f"relatorio_{base_name}.xlsx"
    processar_e_exportar(sys.argv[1], saida)
