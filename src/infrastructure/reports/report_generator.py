# src/report_generator.py

import csv
import logging
import os
import openpyxl
from openpyxl import utils as openpyxl_utils
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple


# Configuração de logger específico para o módulo
logger = logging.getLogger(__name__)

# --- CONSTANTES ---
REGIME_INTERNACAO_MAP = {
    '1': 'Hospitalar',
    '2': 'Hospital-Dia',
    '3': 'Domiciliar'
}

# --- FUNÇÕES AUXILIARES ---

def _formatar_competencia_aaaamm(competencia_aamm_str: Optional[str],
                                 data_emissao_yyyymmdd_str: Optional[str]) -> Optional[str]:
    """Formata a competência no formato AAAAMM."""
    if not competencia_aamm_str:
        return competencia_aamm_str
        
    if len(competencia_aamm_str) == 4 and data_emissao_yyyymmdd_str and len(data_emissao_yyyymmdd_str) == 8:
        try:
            ano_emissao_completo = data_emissao_yyyymmdd_str[0:4]
            seculo_emissao = ano_emissao_completo[0:2]
            ano_competencia_curto = competencia_aamm_str[0:2]
            mes_competencia = competencia_aamm_str[2:4]
            return f"{seculo_emissao}{ano_competencia_curto}{mes_competencia}"
        except Exception:
            return competencia_aamm_str
    elif len(competencia_aamm_str) == 6:
        return competencia_aamm_str
        
    return competencia_aamm_str

def _formatar_data_para_relatorio(data_str_yyyymmdd: Optional[str]) -> str:
    """Formata data YYYYMMDD para DD/MM/YYYY."""
    if not data_str_yyyymmdd or len(data_str_yyyymmdd) != 8:
        return data_str_yyyymmdd or 'N/A'
    
    try:
        dt_obj = datetime.strptime(data_str_yyyymmdd, '%Y%m%d')
        return dt_obj.strftime('%d/%m/%Y')
    except ValueError:
        return data_str_yyyymmdd

def _formatar_valor_para_numero(valor_str: Optional[str]) -> float:
    """Converte string de valor para float."""
    if not valor_str:
        return 0.0
        
    try:
        valor_corrigido_str = str(valor_str).replace(',', '.')
        return float(valor_corrigido_str)
    except (ValueError, TypeError):
        logging.warning(f"Não foi possível converter valor '{valor_str}' para número. Usando 0.0.")
        logger.warning(f"Não foi possível converter valor '{valor_str}' para número. Usando 0.0.")
        return 0.0

def _formatar_unimed_destino(cod_unimed: Optional[str], nome_unimed: Optional[str]) -> str:
    """Formata a descrição da Unimed destino."""
    if not cod_unimed:
        return nome_unimed or 'N/A'
        
    nome_unimed_str = str(nome_unimed)
    if nome_unimed and "não encontrada" not in nome_unimed_str.lower() and "não mapeado" not in nome_unimed_str.lower():
        return f"{cod_unimed} - {nome_unimed_str}"
    else:
        return f"{cod_unimed} - (Nome não localizado)"

# --- FUNÇÕES PRINCIPAIS ---

def gerar_relatorio_distribuicao(plano_distribuicao: Dict[str, Any],
                                 caminho_pasta_distribuicao: str) -> Tuple[bool, Optional[str]]:
    """
    Gera relatório Excel com a distribuição das faturas por auditor.
    
    NOVO: Também salva uma cópia no histórico e auto-importa para o banco de faturas.
    """
    if not plano_distribuicao:
        logger.error("Plano de distribuição está vazio. Relatório Excel não gerado.")
        return False, None

    nome_arquivo_excel = "DISTRIBUIÇÃO.xlsx"
    caminho_completo_excel = os.path.join(caminho_pasta_distribuicao, nome_arquivo_excel)
    
    # Lista para coletar faturas para importação
    faturas_para_importar = []

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        if sheet is None:
            logger.error("Não foi possível obter a planilha ativa do novo workbook Excel.")
            return False, None
        
        sheet.title = "Distribuição Faturas Audit+"

        cabecalhos = [
            "Nº FATURA", "COMP", "UNIMED",
            "EMISSÃO", "VENCIMENTO", "VALOR",
            "AUDITOR"
        ]
        sheet.append(cabecalhos)

        for nome_auditor, dados_auditor in plano_distribuicao.items():
            for fatura_info in dados_auditor.get('faturas', []):
                competencia_original = fatura_info.get('competencia')
                data_emissao_original = fatura_info.get('data_emissao')
                competencia_fmt = _formatar_competencia_aaaamm(competencia_original, data_emissao_original)

                cod_uni_destino = fatura_info.get('codigo_unimed_destino')
                nome_uni_destino = fatura_info.get('nome_unimed_destino')
                unimed_destino_formatada = _formatar_unimed_destino(cod_uni_destino, nome_uni_destino)

                valor_numerico = _formatar_valor_para_numero(fatura_info.get('valor_total_documento'))
                
                linha_dados = [
                    fatura_info.get('numero_fatura', 'N/A'),
                    competencia_fmt,
                    unimed_destino_formatada,
                    _formatar_data_para_relatorio(data_emissao_original),
                    _formatar_data_para_relatorio(fatura_info.get('data_vencimento')),
                    valor_numerico,
                    nome_auditor
                ]
                sheet.append(linha_dados)
                
                cell_valor = sheet.cell(row=sheet.max_row, column=cabecalhos.index("VALOR") + 1)
                cell_valor.number_format = 'R$ #,##0.00'
                
                # Coletar fatura para importação ao banco
                faturas_para_importar.append({
                    'nro_fatura': str(fatura_info.get('numero_fatura', '')),
                    'competencia': competencia_fmt,
                    'unimed_codigo': cod_uni_destino,
                    'unimed_nome': nome_uni_destino,
                    'valor': valor_numerico,
                    'responsavel': nome_auditor,
                    'status': 'PENDENTE'
                })
        
        for col_idx, column in enumerate(sheet.columns, 1):
            max_length = 0
            column_letter = openpyxl_utils.get_column_letter(col_idx)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length > 0 else 12
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Salvar na pasta de distribuição
        workbook.save(filename=caminho_completo_excel)
        logger.info(f"Relatório '{nome_arquivo_excel}' gerado com sucesso em '{caminho_pasta_distribuicao}'.")
        
        # ✅ NOVO: Salvar cópia no histórico com timestamp
        try:
            historico_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), 'data', 'historico_distribuicao')
            os.makedirs(historico_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_historico = f"DISTRIBUICAO_{timestamp}.xlsx"
            caminho_historico = os.path.join(historico_dir, nome_historico)
            
            workbook.save(filename=caminho_historico)
            logger.info(f"📁 Cópia salva no histórico: {caminho_historico}")
        except Exception as e:
            logger.warning(f"Não foi possível salvar cópia no histórico: {e}")
        
        # ✅ NOVO: Auto-importar faturas para o banco de consulta
        try:
            from src.database.fatura_repository import importar_lote
            stats = importar_lote(faturas_para_importar, f"Distribuição {datetime.now().strftime('%d/%m/%Y')}")
            logger.info(f"📊 {stats['criadas']} faturas criadas, {stats['atualizadas']} atualizadas no banco.")
        except Exception as e:
            logger.warning(f"Não foi possível importar faturas para o banco: {e}")
        
        return True, caminho_completo_excel

    except Exception as e:
        logger.exception(f"Falha ao gerar o relatório Excel. Erro: {e}")
        return False, None

def gerar_csv_internacao(guias_relevantes: List[Dict[str, Any]],
                         output_folder: str) -> bool:
    """Gera CSV com guias de internação relevantes."""
    if not guias_relevantes:
        logger.warning("Nenhuma guia de internação relevante fornecida. CSV não gerado.")
        return True

    output_filename = "Guias de Internação Relevantes.csv"
    output_path = os.path.join(output_folder, output_filename)

    headers = [
        "Fatura Pai", "Nº Guia Internação", "Código Beneficiário",
        "Nome Beneficiário", "Regime de Internação",
        "Valor p/ Filtro (R$)", "Valor Real Total (R$)"
    ]

    logger.info(f"Gerando CSV de guias de internação: {output_path}")

    try:
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerow(headers)

            for guia in guias_relevantes:
                valor_filtro_str = f"{guia.get('valor_filtro', 0.0):.2f}".replace('.', ',')
                valor_real_str = f"{guia.get('valor_total_real', 0.0):.2f}".replace('.', ',')

                regime_cod = guia.get('regime_internacao', '')
                regime_desc = REGIME_INTERNACAO_MAP.get(regime_cod, regime_cod)

                writer.writerow([
                    guia.get('fatura_pai', ''),
                    guia.get('numero_guia', ''),
                    guia.get('codigo_beneficiario', ''),
                    guia.get('nome_beneficiario', ''),
                    regime_desc,
                    valor_filtro_str,
                    valor_real_str
                ])

        logger.info(f"CSV gerado com sucesso: {len(guias_relevantes)} guias")
        return True
        
    except IOError as e:
        logger.exception(f"Erro de E/S ao escrever CSV em {output_path}: {e}")
        return False
    except Exception as e:
        logger.exception(f"Erro inesperado ao gerar CSV: {e}")
        return False

def gerar_csv_alertas_internacao_curta(guias_para_sinalizar: List[Dict[str, Any]],
                                       output_folder: str) -> Tuple[bool, str]:
    """Gera CSV com alertas de internações curtas para revisão."""
    if not guias_para_sinalizar:
        return True, "Nenhuma guia com inconsistência de internação curta encontrada."

    output_filename = "ALERTA_Internacoes_Curtas_a_Revisar.csv"
    output_path = os.path.join(output_folder, output_filename)
    
    headers = [
        "Arquivo Origem", "Nº Guia Internação", "Nome Beneficiário",
        "Regime Informado", "Duração", "Motivo do Alerta"
    ]
    
    try:
        # Criar diretório se não existir
        os.makedirs(output_folder, exist_ok=True)
        
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerow(headers)

            linhas_escritas = 0
            for guia in guias_para_sinalizar:
                try:
                    linha = [
                        guia.get('arquivo_origem', ''),
                        guia.get('numero_guia', ''),
                        guia.get('nome_beneficiario', ''),
                        guia.get('regime_informado', ''),
                        guia.get('duracao', ''),
                        guia.get('motivo', '')
                    ]
                    writer.writerow(linha)
                    linhas_escritas += 1
                    
                except Exception as e:
                    logger.warning(f"Erro ao escrever guia no CSV: {e}")
                    continue

        if linhas_escritas == 0:
            try:
                os.remove(output_path)  # Remove arquivo vazio
            except:
                pass
            return False, "Nenhum alerta pôde ser escrito no CSV."

        logger.info(f"CSV de alertas gerado com sucesso: {linhas_escritas} alertas")
        return True, f"Relatório de alertas gerado com {linhas_escritas} registros: {output_path}"
        
    except IOError as e:
        logger.exception(f"Erro de E/S ao gerar CSV de alertas em {output_path}: {e}")
        return False, f"Erro de acesso ao arquivo: {e}"
    except Exception as e:
        logger.exception(f"Erro inesperado ao gerar CSV de alertas: {e}")
        return False, f"Erro inesperado: {e}"