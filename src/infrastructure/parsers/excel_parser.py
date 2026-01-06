"""
Glox - Parser de Excel para Importação

Parseia arquivos Excel do BI e converte para formato de fatura.
Suporta: A500 enviados, Distribuição de Faturas, Faturas Emitidas
"""

from typing import List, Dict, Optional
from datetime import datetime
from openpyxl import load_workbook
import re


def parse_a500_enviados(filepath: str) -> List[Dict]:
    """
    Parseia arquivo A500 enviados.xlsx
    
    Estrutura esperada (linha 3+):
    - N. Doc 1, Etiqueta, N. Doc 2, Responsavel, ..., Valor Total Fatura
    
    Returns:
        Lista de dicionários com dados das faturas
    """
    faturas = []
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        rows = list(ws.rows)
        if len(rows) < 4:
            return faturas
        
        # Linha 3 tem os headers
        headers = [str(cell.value).strip() if cell.value else "" for cell in rows[2]]
        
        # Mapear índices das colunas importantes
        col_map = {}
        for i, h in enumerate(headers):
            h_lower = h.lower()
            if 'doc 1' in h_lower or 'doc1' in h_lower:
                col_map['nro_fatura'] = i
            elif 'responsavel' in h_lower or 'responsável' in h_lower:
                col_map['responsavel'] = i
            elif 'valor total' in h_lower:
                col_map['valor'] = i
            elif 'unimed' in h_lower:
                col_map['unimed'] = i
        
        # Processar dados (linha 4+)
        for row in rows[3:]:
            cells = [cell.value for cell in row]
            
            nro_fatura = None
            if 'nro_fatura' in col_map:
                nro_fatura = str(cells[col_map['nro_fatura']] or "").strip()
            
            if not nro_fatura or nro_fatura == 'None':
                continue
            
            fatura = {
                'nro_fatura': nro_fatura,
                'status': 'ENVIADA',  # A500 = já enviados
                'responsavel': str(cells[col_map.get('responsavel', 0)] or "").strip() if 'responsavel' in col_map else "",
            }
            
            if 'valor' in col_map:
                val = cells[col_map['valor']]
                fatura['valor'] = float(val) if val else 0.0
            
            faturas.append(fatura)
        
        wb.close()
        
    except Exception as e:
        print(f"Erro ao parsear A500: {e}")
    
    return faturas


def parse_distribuicao_faturas(filepath: str) -> List[Dict]:
    """
    Parseia arquivo Distribuição de Faturas Intercâmbio.xlsx
    
    Estrutura esperada:
    - NRO_FATURA, DAT_COMPET, DESCRICAO, EMISSAO, ...
    
    Returns:
        Lista de dicionários com dados das faturas
    """
    faturas = []
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        rows = list(ws.rows)
        if len(rows) < 2:
            return faturas
        
        # Primeira linha = headers
        headers = [str(cell.value).strip().upper() if cell.value else "" for cell in rows[0]]
        
        # Mapear colunas
        col_map = {}
        for i, h in enumerate(headers):
            if 'NRO_FATURA' in h or 'FATURA' in h:
                col_map['nro_fatura'] = i
            elif 'COMPET' in h:
                col_map['competencia'] = i
            elif 'VALOR' in h:
                col_map['valor'] = i
            elif 'DESCRICAO' in h:
                col_map['descricao'] = i
        
        # Processar dados
        for row in rows[1:]:
            cells = [cell.value for cell in row]
            
            nro_fatura = None
            if 'nro_fatura' in col_map:
                nro_fatura = str(cells[col_map['nro_fatura']] or "").strip()
            
            if not nro_fatura or nro_fatura == 'None':
                continue
            
            fatura = {
                'nro_fatura': nro_fatura,
                'status': 'PENDENTE',  # Distribuição = ainda não enviado
            }
            
            if 'competencia' in col_map:
                comp = cells[col_map['competencia']]
                if comp:
                    # Formatar como MM/YYYY
                    if isinstance(comp, datetime):
                        fatura['competencia'] = comp.strftime("%m/%Y")
                    else:
                        fatura['competencia'] = str(comp)[:7]
            
            if 'valor' in col_map:
                val = cells[col_map['valor']]
                fatura['valor'] = float(val) if val else 0.0
            
            # Extrair Unimed da descrição
            if 'descricao' in col_map:
                desc = str(cells[col_map['descricao']] or "")
                match = re.search(r'UNIMED[:\s]+(.+?)(?:\s*-|$)', desc, re.IGNORECASE)
                if match:
                    fatura['unimed_nome'] = match.group(1).strip()
            
            faturas.append(fatura)
        
        wb.close()
        
    except Exception as e:
        print(f"Erro ao parsear Distribuição: {e}")
    
    return faturas


def parse_faturas_emitidas(filepath: str) -> List[Dict]:
    """
    Parseia arquivo faturas_emitidas.xlsx
    
    Returns:
        Lista de dicionários com dados das faturas
    """
    faturas = []
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        rows = list(ws.rows)
        if len(rows) < 2:
            return faturas
        
        headers = [str(cell.value).strip().upper() if cell.value else "" for cell in rows[0]]
        
        # Mapear colunas
        col_map = {}
        for i, h in enumerate(headers):
            h_clean = h.replace('_', '').replace(' ', '')
            if 'FATURA' in h_clean or 'NRO' in h_clean:
                col_map['nro_fatura'] = i
            elif 'VALOR' in h_clean:
                col_map['valor'] = i
            elif 'STATUS' in h_clean:
                col_map['status'] = i
            elif 'UNIMED' in h_clean:
                col_map['unimed'] = i
        
        for row in rows[1:]:
            cells = [cell.value for cell in row]
            
            nro_fatura = None
            if 'nro_fatura' in col_map:
                nro_fatura = str(cells[col_map['nro_fatura']] or "").strip()
            
            if not nro_fatura or nro_fatura == 'None':
                continue
            
            fatura = {
                'nro_fatura': nro_fatura,
                'status': 'PENDENTE',
            }
            
            if 'valor' in col_map:
                val = cells[col_map['valor']]
                fatura['valor'] = float(val) if val else 0.0
            
            if 'status' in col_map:
                st = str(cells[col_map['status']] or "").upper()
                if 'ENVI' in st:
                    fatura['status'] = 'ENVIADA'
                elif 'CANCEL' in st:
                    fatura['status'] = 'CANCELADA'
                elif 'GLOS' in st:
                    fatura['status'] = 'GLOSADA'
            
            faturas.append(fatura)
        
        wb.close()
        
    except Exception as e:
        print(f"Erro ao parsear Faturas Emitidas: {e}")
    
    return faturas


def detectar_tipo_arquivo(filepath: str) -> Optional[str]:
    """
    Detecta o tipo de arquivo Excel baseado no nome ou conteúdo.
    
    Returns:
        'A500' | 'DISTRIBUICAO' | 'EMITIDAS' | None
    """
    filename = filepath.lower()
    
    if 'a500' in filename:
        return 'A500'
    elif 'distribuic' in filename or 'distrib' in filename:
        return 'DISTRIBUICAO'
    elif 'emitid' in filename:
        return 'EMITIDAS'
    
    # Tentar detectar pelo conteúdo
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        first_row = [str(cell.value or "").upper() for cell in list(ws.rows)[0]]
        wb.close()
        
        text = " ".join(first_row)
        if 'NRO_FATURA' in text or 'DAT_COMPET' in text:
            return 'DISTRIBUICAO'
        elif 'DOC 1' in text or 'ETIQUETA' in text:
            return 'A500'
            
    except:
        pass
    
    return 'EMITIDAS'  # Default


def parse_arquivo(filepath: str) -> List[Dict]:
    """
    Parseia automaticamente qualquer arquivo Excel suportado.
    
    Args:
        filepath: Caminho do arquivo Excel
    
    Returns:
        Lista de faturas parseadas
    """
    tipo = detectar_tipo_arquivo(filepath)
    
    if tipo == 'A500':
        return parse_a500_enviados(filepath)
    elif tipo == 'DISTRIBUICAO':
        return parse_distribuicao_faturas(filepath)
    else:
        return parse_faturas_emitidas(filepath)
